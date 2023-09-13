import logging
import os

import pandas as pd
import datetime
import numpy as np
from lazyutils.config.Configuration import ConfigFromEnv
from sqlalchemy.orm import Session

from contextualization.BankDataMerging import BankDataMerging
from contextualization.XPInvestmentRentability import process_tesouro_direto, process_renda_fixa, \
    process_fundos_imobiliarios, process_compromissadas, process_previdencia_privada, process_fundos_de_investimento, \
    process_coe, process_acoes
from models.BankStatements import Bank
from models.Investments import InvestmentType
from openpyxl import load_workbook

from models.base import engine

from contextualization import InvestmentsFactory


class XPInvestmentDataMerging(BankDataMerging):
    SKIP_ROWS = 4
    FILE_DATE_COLUMN = 6
    FILE_DATE_ROW = 1

    def __init__(self):
        super().__init__()
        self.category_map = None
        self.bank = None

    def process_investment_rows(self, rows, investment_type: InvestmentType, file_date) -> list:
        pd.DataFrame(rows)

        if investment_type is InvestmentType.TESOURO_DIRETO:
            return process_tesouro_direto(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.RENDA_FIXA:
            return process_renda_fixa(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.FUNDOS_IMOBILIARIOS:
            return process_fundos_imobiliarios(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.COMPROMISSADAS:
            return process_compromissadas(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.PREVIDENCIA_PRIVADA:
            return process_previdencia_privada(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.FUNDOS_DE_INVESTIMENTO:
            return process_fundos_de_investimento(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.COE:
            return process_coe(rows, file_date, self.bank.id)
        elif investment_type is InvestmentType.ACOES:
            return process_acoes(rows, file_date, self.bank.id)

        return []

    def read_excel_data(self, filepath):
        wb = load_workbook(filename=filepath)
        ws = wb['Sua carteira']
        file_date = datetime.datetime.strptime(
            ws.cell(row=self.FILE_DATE_ROW, column=self.FILE_DATE_COLUMN)
            .value.split(' | ')[1], '%d/%m/%Y, %H:%M'
        )

        df = pd.read_excel(filepath, skiprows=self.SKIP_ROWS, header=None, engine='openpyxl')
        df.replace(['', ' '], np.nan, inplace=True)
        df['is_blank'] = df.isnull().all(axis=1)

        wb.close()
        del wb

        return df, file_date

    def process_data(self, df, file_date):
        """
        This function is responsible for processing the data from the excel file.
        It will read the dataframe and process the data.
        The data is processed in a way that the investments are grouped by investment type.
        """
        current_type = None
        investment_rows = []
        investments_processed = []
        lines_processed = 0

        for index, row in df.iterrows():
            if row['is_blank']:
                continue

            try:
                if row[0] in self.category_map:
                    if current_type:
                        investments_processed.extend(
                            self.process_investment_rows(investment_rows, current_type, file_date)
                        )
                        lines_processed += 1

                    current_type = self.category_map[row[0]]
                    investment_rows = []
                    continue

                investment_rows.append(row)
            except TypeError as te:
                logging.error(f'Erro ao processar linha: {row} - {str(te)}')
            except ValueError as ve:
                logging.error(f'Erro ao processar linha: {row} - {str(ve)}')

        return investments_processed, lines_processed

    def process_investment_rentability_from_excel_file(self, excel_file_path, session: Session):

        self.category_map = InvestmentsFactory.investmenttype_map()

        df, file_date = self.read_excel_data(excel_file_path)
        investments_processed, lines_processed = self.process_data(df, file_date)

        investmentsfactory = InvestmentsFactory.InvestmentsFactory()

        for investmenttuple in investments_processed:
            investment_key, rentability_key = investmentsfactory.create_key(investmenttuple[0], investmenttuple[1])

            investment, investment_rentability = investmentsfactory.process_investment(
                investmenttuple[0], investmenttuple[1], investment_key, session)

            investmentsfactory.process_investment_rentability(investment, investment_rentability,
                                                              rentability_key, session)

        session.commit()

        return file_date, lines_processed

    def merge_bank_statement_data(self, file_folder):
        files = [file for file in os.listdir(file_folder)
                 if file.startswith("PosicaoDetalhada_") and file.endswith(".xlsx")]

        with Session(engine) as session:
            self.bank = session.query(Bank).filter_by(name='XP').first()

            for file in files:

                if self.is_file_processed(file, session):
                    logging.info(f"Skipping previously processed file: {file}")
                    continue

                file_path = os.path.join(file_folder, file)
                fdate, lines_processed = self.process_investment_rentability_from_excel_file(file_path, session)

                # Update the list of processed files
                self.update_processed_file(file, 'Processed', lines_processed, session)

                # Move the processed file to the 'investmentprocessedfolder' folder
                self.move_file_to_processed_folder(file_path, 'investmentprocessedfolder')

            InvestmentsFactory.update_retability(session)

            session.commit()


if __name__ == '__main__':
    config = ConfigFromEnv()

    xpinvest = XPInvestmentDataMerging()
    xpinvest.merge_bank_statement_data(config['wallet']['investmentfolder'])
